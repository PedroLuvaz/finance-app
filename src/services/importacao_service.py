"""
Serviço de importação de faturas.
Suporta importação de arquivos XLS/XLSX de diferentes bancos.
"""

from dataclasses import dataclass
from datetime import datetime, date
from typing import List, Optional, Dict, Any, Protocol
from pathlib import Path
from abc import ABC, abstractmethod
import re

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    openpyxl = None

try:
    import pandas as pd
except ImportError:
    pd = None

from src.services.conta_service import ContaService, DadosConta, ResultadoOperacao


@dataclass
class TransacaoImportada:
    """Representa uma transação importada de um arquivo."""
    descricao: str
    valor: float
    data: datetime
    parcela_atual: int = 1
    total_parcelas: int = 1
    categoria_sugerida: Optional[str] = None
    
    @property
    def eh_parcelada(self) -> bool:
        return self.total_parcelas > 1


class ParserFatura(ABC):
    """Interface base para parsers de fatura."""
    
    @property
    @abstractmethod
    def nome_banco(self) -> str:
        """Nome do banco que este parser suporta."""
        pass
    
    @property
    @abstractmethod
    def extensoes_suportadas(self) -> List[str]:
        """Extensões de arquivo suportadas."""
        pass
    
    @abstractmethod
    def pode_processar(self, caminho: Path) -> bool:
        """Verifica se este parser pode processar o arquivo."""
        pass
    
    @abstractmethod
    def processar(self, caminho: Path) -> List[TransacaoImportada]:
        """Processa o arquivo e retorna as transações."""
        pass
    
    def _extrair_parcelas(self, descricao: str) -> tuple:
        """Extrai informações de parcelas da descrição."""
        # Padrão Nubank: "- Parcela X/Y" ou "Parcela X/Y"
        match = re.search(r'-?\s*[Pp]arcela\s+(\d+)/(\d+)', descricao)
        if match:
            parcela_atual = int(match.group(1))
            total_parcelas = int(match.group(2))
            # Remove a informação de parcela da descrição
            descricao_limpa = re.sub(r'\s*-?\s*[Pp]arcela\s+\d+/\d+\s*', '', descricao).strip()
            return descricao_limpa, parcela_atual, total_parcelas
        
        # Padrão simples: "X/Y" no final
        match = re.search(r'\s+(\d+)/(\d+)\s*$', descricao)
        if match:
            parcela_atual = int(match.group(1))
            total_parcelas = int(match.group(2))
            descricao_limpa = re.sub(r'\s+\d+/\d+\s*$', '', descricao).strip()
            return descricao_limpa, parcela_atual, total_parcelas
        
        return descricao, 1, 1


class NubankParser(ParserFatura):
    """Parser para faturas do Nubank."""
    
    @property
    def nome_banco(self) -> str:
        return "Nubank"
    
    @property
    def extensoes_suportadas(self) -> List[str]:
        return ['.xlsx', '.xls', '.csv']
    
    def pode_processar(self, caminho: Path) -> bool:
        """Verifica se é um arquivo do Nubank."""
        nome = caminho.name.lower()
        return 'nubank' in nome or 'nu' in nome
    
    def processar(self, caminho: Path) -> List[TransacaoImportada]:
        """Processa fatura do Nubank."""
        transacoes = []
        
        if caminho.suffix.lower() == '.csv':
            transacoes = self._processar_csv(caminho)
        else:
            transacoes = self._processar_excel(caminho)
        
        return transacoes
    
    def _processar_csv(self, caminho: Path) -> List[TransacaoImportada]:
        """Processa arquivo CSV do Nubank."""
        transacoes = []
        
        try:
            # Tentar diferentes encodings
            encodings = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']
            content = None
            
            for enc in encodings:
                try:
                    with open(caminho, 'r', encoding=enc) as f:
                        content = f.read()
                        break
                except UnicodeDecodeError:
                    continue
            
            if not content:
                print("Não foi possível ler o arquivo com nenhum encoding")
                return transacoes
            
            import csv
            from io import StringIO
            
            reader = csv.DictReader(StringIO(content))
            
            # Debug: mostrar colunas encontradas
            print(f"Colunas encontradas: {reader.fieldnames}")
            
            for row in reader:
                # Colunas típicas do Nubank CSV: date, title, amount
                data_str = row.get('date', row.get('Date', row.get('Data', '')))
                descricao = row.get('title', row.get('Title', row.get('Descrição', row.get('Título', ''))))
                valor_str = row.get('amount', row.get('Amount', row.get('Valor', '0')))
                
                if not descricao:
                    continue
                
                # Parse da data
                try:
                    data = self._parse_data(data_str)
                except:
                    data = datetime.now()
                
                # Parse do valor
                valor = self._parse_valor(valor_str)
                
                # Ignorar valores negativos (créditos/pagamentos recebidos)
                if valor < 0:
                    continue
                
                # Extrair parcelas
                descricao_limpa, parcela, total = self._extrair_parcelas(descricao)
                
                transacoes.append(TransacaoImportada(
                    descricao=descricao_limpa,
                    valor=valor,
                    data=data,
                    parcela_atual=parcela,
                    total_parcelas=total
                ))
                
            print(f"Total de transações encontradas: {len(transacoes)}")
            
        except Exception as e:
            print(f"Erro ao processar CSV: {e}")
            import traceback
            traceback.print_exc()
        
        return transacoes
    
    def _processar_excel(self, caminho: Path) -> List[TransacaoImportada]:
        """Processa arquivo Excel do Nubank."""
        transacoes = []
        
        if pd is not None:
            return self._processar_com_pandas(caminho)
        elif openpyxl is not None:
            return self._processar_com_openpyxl(caminho)
        
        return transacoes
    
    def _processar_com_pandas(self, caminho: Path) -> List[TransacaoImportada]:
        """Processa usando pandas."""
        transacoes = []
        
        try:
            df = pd.read_excel(caminho)
            
            # Detectar colunas
            col_data = self._encontrar_coluna(df.columns, ['date', 'data', 'dt'])
            col_descricao = self._encontrar_coluna(df.columns, ['title', 'descrição', 'descricao', 'titulo', 'título'])
            col_valor = self._encontrar_coluna(df.columns, ['amount', 'valor', 'value'])
            
            for _, row in df.iterrows():
                try:
                    descricao = str(row[col_descricao]) if col_descricao else ''
                    if not descricao or descricao == 'nan':
                        continue
                    
                    # Data
                    if col_data:
                        data_val = row[col_data]
                        if isinstance(data_val, datetime):
                            data = data_val
                        else:
                            data = self._parse_data(str(data_val))
                    else:
                        data = datetime.now()
                    
                    # Valor
                    if col_valor:
                        valor = self._parse_valor(row[col_valor])
                    else:
                        valor = 0.0
                    
                    # Ignorar valores negativos (créditos/pagamentos recebidos)
                    if valor < 0:
                        continue
                    
                    # Extrair parcelas
                    descricao_limpa, parcela, total = self._extrair_parcelas(descricao)
                    
                    transacoes.append(TransacaoImportada(
                        descricao=descricao_limpa,
                        valor=abs(valor),
                        data=data,
                        parcela_atual=parcela,
                        total_parcelas=total
                    ))
                except Exception as e:
                    continue
                    
        except Exception as e:
            print(f"Erro ao processar com pandas: {e}")
        
        return transacoes
    
    def _processar_com_openpyxl(self, caminho: Path) -> List[TransacaoImportada]:
        """Processa usando openpyxl."""
        transacoes = []
        
        try:
            wb = load_workbook(caminho)
            ws = wb.active
            
            # Encontrar cabeçalhos na primeira linha
            headers = {}
            for col, cell in enumerate(ws[1], 1):
                if cell.value:
                    headers[str(cell.value).lower()] = col
            
            # Mapear colunas
            col_data = headers.get('date', headers.get('data', None))
            col_descricao = headers.get('title', headers.get('descrição', headers.get('descricao', None)))
            col_valor = headers.get('amount', headers.get('valor', None))
            
            # Processar linhas (começando da 2ª)
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    descricao = str(row[col_descricao - 1]) if col_descricao else ''
                    if not descricao or descricao == 'None':
                        continue
                    
                    # Data
                    if col_data:
                        data_val = row[col_data - 1]
                        if isinstance(data_val, datetime):
                            data = data_val
                        else:
                            data = self._parse_data(str(data_val))
                    else:
                        data = datetime.now()
                    
                    # Valor
                    if col_valor:
                        valor = self._parse_valor(row[col_valor - 1])
                    else:
                        valor = 0.0
                    
                    # Ignorar valores negativos (créditos/pagamentos recebidos)
                    if valor < 0:
                        continue
                    
                    # Extrair parcelas
                    descricao_limpa, parcela, total = self._extrair_parcelas(descricao)
                    
                    transacoes.append(TransacaoImportada(
                        descricao=descricao_limpa,
                        valor=abs(valor),
                        data=data,
                        parcela_atual=parcela,
                        total_parcelas=total
                    ))
                except Exception:
                    continue
                    
        except Exception as e:
            print(f"Erro ao processar com openpyxl: {e}")
        
        return transacoes
    
    def _encontrar_coluna(self, colunas, opcoes: List[str]) -> Optional[str]:
        """Encontra uma coluna por nome."""
        for col in colunas:
            if str(col).lower() in opcoes:
                return col
        return None
    
    def _parse_data(self, valor: str) -> datetime:
        """Converte string para datetime."""
        formatos = ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d']
        for fmt in formatos:
            try:
                return datetime.strptime(valor.strip(), fmt)
            except:
                continue
        return datetime.now()
    
    def _parse_valor(self, valor) -> float:
        """Converte valor para float."""
        if isinstance(valor, (int, float)):
            return float(valor)
        
        valor_str = str(valor).strip()
        # Remove símbolos monetários e espaços
        valor_str = re.sub(r'[R$\s]', '', valor_str)
        # Trata formato brasileiro (1.234,56) vs americano (1,234.56)
        if ',' in valor_str and '.' in valor_str:
            # Formato brasileiro
            valor_str = valor_str.replace('.', '').replace(',', '.')
        elif ',' in valor_str:
            valor_str = valor_str.replace(',', '.')
        
        try:
            return float(valor_str)
        except:
            return 0.0


class InterParser(ParserFatura):
    """Parser para faturas do Banco Inter."""
    
    @property
    def nome_banco(self) -> str:
        return "Banco Inter"
    
    @property
    def extensoes_suportadas(self) -> List[str]:
        return ['.xlsx', '.xls', '.csv']
    
    def pode_processar(self, caminho: Path) -> bool:
        nome = caminho.name.lower()
        return 'inter' in nome
    
    def processar(self, caminho: Path) -> List[TransacaoImportada]:
        """Processa fatura do Inter - estrutura similar ao Nubank."""
        # Reutiliza a lógica do Nubank pois o formato é similar
        nubank_parser = NubankParser()
        return nubank_parser.processar(caminho)


class GenericoParser(ParserFatura):
    """Parser genérico para qualquer banco."""
    
    @property
    def nome_banco(self) -> str:
        return "Genérico"
    
    @property
    def extensoes_suportadas(self) -> List[str]:
        return ['.xlsx', '.xls', '.csv']
    
    def pode_processar(self, caminho: Path) -> bool:
        """Aceita qualquer arquivo com extensão suportada."""
        return caminho.suffix.lower() in self.extensoes_suportadas
    
    def processar(self, caminho: Path) -> List[TransacaoImportada]:
        """Tenta processar de forma genérica."""
        # Usa a mesma lógica do Nubank como base
        nubank_parser = NubankParser()
        return nubank_parser.processar(caminho)


class ImportacaoService:
    """Serviço para importação de faturas."""
    
    def __init__(self):
        self.conta_service = ContaService()
        self.parsers: List[ParserFatura] = [
            NubankParser(),
            InterParser(),
            GenericoParser(),  # Sempre por último
        ]
    
    def listar_bancos_suportados(self) -> List[str]:
        """Retorna lista de bancos suportados."""
        return [p.nome_banco for p in self.parsers if p.nome_banco != "Genérico"]
    
    def verificar_dependencias(self) -> Dict[str, bool]:
        """Verifica se as dependências estão instaladas."""
        return {
            'openpyxl': openpyxl is not None,
            'pandas': pd is not None,
        }
    
    def importar_arquivo(
        self, 
        caminho: str,
        banco: Optional[str] = None
    ) -> ResultadoOperacao:
        """
        Importa transações de um arquivo.
        
        Args:
            caminho: Caminho do arquivo
            banco: Nome do banco (opcional, tenta detectar automaticamente)
            
        Returns:
            ResultadoOperacao com lista de TransacaoImportada em dados
        """
        path = Path(caminho)
        
        if not path.exists():
            return ResultadoOperacao(sucesso=False, mensagem="Arquivo não encontrado")
        
        if path.suffix.lower() not in ['.xlsx', '.xls', '.csv']:
            return ResultadoOperacao(sucesso=False, mensagem="Formato de arquivo não suportado. Use .xlsx, .xls ou .csv")
        
        # Encontrar parser apropriado
        parser = self._encontrar_parser(path, banco)
        
        if not parser:
            return ResultadoOperacao(sucesso=False, mensagem="Não foi possível identificar o formato do arquivo")
        
        try:
            transacoes = parser.processar(path)
            
            if not transacoes:
                return ResultadoOperacao(sucesso=False, mensagem="Nenhuma transação encontrada no arquivo")
            
            return ResultadoOperacao(sucesso=True, dados=transacoes)
            
        except Exception as e:
            return ResultadoOperacao(sucesso=False, mensagem=f"Erro ao processar arquivo: {str(e)}")
    
    def _encontrar_parser(self, caminho: Path, banco: Optional[str]) -> Optional[ParserFatura]:
        """Encontra o parser apropriado para o arquivo."""
        if banco:
            for parser in self.parsers:
                if parser.nome_banco.lower() == banco.lower():
                    return parser
        
        # Tenta detectar automaticamente
        for parser in self.parsers:
            if parser.pode_processar(caminho):
                return parser
        
        return None
    
    def salvar_transacoes(
        self,
        transacoes: List[TransacaoImportada],
        categoria_id: int,
        divisoes: Optional[List[Dict]] = None,
        gerar_parcelas_futuras: bool = False
    ) -> ResultadoOperacao:
        """
        Salva as transações importadas no banco de dados.
        
        Args:
            transacoes: Lista de transações a salvar
            categoria_id: ID da categoria
            divisoes: Lista de divisões (opcional)
            gerar_parcelas_futuras: Se deve gerar parcelas futuras
            
        Returns:
            ResultadoOperacao da operação
        """
        salvas = 0
        erros = []
        
        for transacao in transacoes:
            try:
                # O valor da transação importada já é o valor da parcela/fatura
                # Não devemos multiplicar pelo total de parcelas, pois o ContaService
                # espera o valor da parcela individual para contas parceladas.
                
                # Só geramos parcelas futuras se o usuário solicitou E a transação é parcelada
                gerar_futuras = gerar_parcelas_futuras and transacao.eh_parcelada
                
                # Converter data para string no formato esperado (YYYY-MM-DD)
                data_vencimento_str = None
                if isinstance(transacao.data, (datetime, date)):
                    data_vencimento_str = transacao.data.strftime('%Y-%m-%d')
                else:
                    # Tenta usar como string ou pega data atual se falhar
                    data_vencimento_str = str(transacao.data) if transacao.data else datetime.now().strftime('%Y-%m-%d')
                
                dados = DadosConta(
                    descricao=transacao.descricao,
                    valor_total=transacao.valor,
                    parcela_atual=transacao.parcela_atual,
                    total_parcelas=transacao.total_parcelas,
                    data_vencimento=data_vencimento_str,
                    categoria_id=categoria_id,
                    gerar_parcelas_futuras=gerar_futuras,
                    divisoes=divisoes or []
                )
                
                resultado = self.conta_service.criar_conta(dados)
                
                if resultado.sucesso:
                    salvas += 1
                else:
                    erros.append(f"{transacao.descricao}: {resultado.mensagem}")
                    
            except Exception as e:
                erros.append(f"{transacao.descricao}: {str(e)}")
        
        if salvas == 0:
            return ResultadoOperacao(sucesso=False, mensagem=f"Nenhuma transação salva. Erros: {'; '.join(erros)}")
        
        mensagem = f"{salvas} transação(ões) importada(s) com sucesso!"
        if erros:
            mensagem += f" ({len(erros)} erro(s))"
        
        return ResultadoOperacao(sucesso=True, mensagem=mensagem)
